#!/usr/bin/env python3
"""Validate a cleaned and semantically audited review-tagging workbook."""

from __future__ import annotations

import argparse
import json
import sys
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook


REQUIRED_DETAIL = {
    "Review ID",
    "主要问题类型",
    "主要原因",
    "一级问题类型",
    "二级具体原因",
    "语义核对结果",
    "修正说明",
    "核对方式",
}
VALID_AUDIT_RESULTS = {"已核对-匹配", "已核对-已修正"}
PLACEHOLDERS = {"其他", "未覆盖", "待分类", "待确认"}


def find_header(sheet, required: set[str], max_rows: int = 20) -> tuple[int, list[str]]:
    for row_number, values in enumerate(
        sheet.iter_rows(min_row=1, max_row=min(max_rows, sheet.max_row), values_only=True), start=1
    ):
        headers = [str(value).strip() if value is not None else "" for value in values]
        if required.issubset(set(headers)):
            return row_number, headers
    raise ValueError(f"{sheet.title}未找到必需表头: {sorted(required)}")


def rows_as_dicts(sheet, header_row: int, headers: list[str]):
    unique_headers = []
    for index, header in enumerate(headers, start=1):
        unique_headers.append(header or f"__blank_{index}")
    for values in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        row = dict(zip(unique_headers, values))
        if row.get("Review ID") not in (None, ""):
            yield row


def split_labels(value) -> set[str]:
    if value in (None, ""):
        return set()
    normalized = str(value).replace("|", "；").replace(",", "；")
    return {part.strip() for part in normalized.split("；") if part.strip()}


def validate(path: Path, detail_sheet: str, dictionary_sheet: str) -> dict:
    workbook = load_workbook(path, read_only=True, data_only=True)
    errors: list[str] = []
    warnings: list[str] = []

    if detail_sheet not in workbook.sheetnames:
        raise ValueError(f"缺少工作表: {detail_sheet}")
    if dictionary_sheet not in workbook.sheetnames:
        raise ValueError(f"缺少工作表: {dictionary_sheet}")

    detail = workbook[detail_sheet]
    detail_header_row, detail_headers = find_header(detail, REQUIRED_DETAIL)
    detail_rows = list(rows_as_dicts(detail, detail_header_row, detail_headers))

    dictionary = workbook[dictionary_sheet]
    dictionary_required = {"一级问题类型", "二级具体原因"}
    dictionary_header_row, dictionary_headers = find_header(dictionary, dictionary_required)
    dictionary_rows = list(rows_as_dicts_with_key(dictionary, dictionary_header_row, dictionary_headers, "一级问题类型"))
    allowed_pairs = {
        (str(row["一级问题类型"]).strip(), str(row["二级具体原因"]).strip())
        for row in dictionary_rows
        if row.get("一级问题类型") and row.get("二级具体原因")
    }

    review_ids = [str(row["Review ID"]).strip() for row in detail_rows]
    duplicate_ids = [key for key, count in Counter(review_ids).items() if count > 1]
    if duplicate_ids:
        errors.append(f"Review ID重复: {duplicate_ids[:10]}")

    audit_counts = Counter()
    confidence_counts = Counter()
    corrected = 0
    information_insufficient = 0
    unknown_pairs = Counter()

    for row in detail_rows:
        review_id = str(row["Review ID"]).strip()
        primary_type = str(row.get("主要问题类型") or "").strip()
        primary_reason = str(row.get("主要原因") or "").strip()
        audit_result = str(row.get("语义核对结果") or "").strip()
        correction_note = str(row.get("修正说明") or "").strip()

        if not primary_type or not primary_reason:
            errors.append(f"{review_id}: 主要问题类型或主要原因为空")
        if any(label in PLACEHOLDERS for label in split_labels(row.get("一级问题类型")) | split_labels(row.get("二级具体原因"))):
            errors.append(f"{review_id}: 存在未解决占位标签")
        if audit_result not in VALID_AUDIT_RESULTS:
            errors.append(f"{review_id}: 语义核对结果无效或为空")
        if audit_result == "已核对-已修正":
            corrected += 1
            if not correction_note:
                errors.append(f"{review_id}: 已修正但缺少修正说明")
        if (primary_type, primary_reason) not in allowed_pairs:
            unknown_pairs[(primary_type, primary_reason)] += 1
        if primary_type == "信息不足":
            information_insufficient += 1

        audit_counts[audit_result] += 1
        confidence = str(row.get("标签置信等级") or row.get("核对置信等级") or "未提供").strip()
        confidence_counts[confidence] += 1

        for optional_field in ("使用场景", "适用人群"):
            if optional_field in detail_headers and not str(row.get(optional_field) or "").strip():
                warnings.append(f"{review_id}: {optional_field}为空，应使用未提及/无法判断")

    if unknown_pairs:
        errors.append(f"主要类型/原因未进入标签字典: {dict(unknown_pairs)}")

    result = {
        "status": "pass" if not errors else "fail",
        "workbook": str(path.resolve()),
        "metrics": {
            "detail_rows": len(detail_rows),
            "unique_review_ids": len(set(review_ids)),
            "audit_counts": dict(audit_counts),
            "corrected_rows": corrected,
            "information_insufficient_rows": information_insufficient,
            "confidence_counts": dict(confidence_counts),
            "dictionary_pairs": len(allowed_pairs),
        },
        "errors": errors,
        "warnings": warnings[:100],
    }
    return result


def rows_as_dicts_with_key(sheet, header_row: int, headers: list[str], key: str):
    unique_headers = []
    for index, header in enumerate(headers, start=1):
        unique_headers.append(header or f"__blank_{index}")
    for values in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        row = dict(zip(unique_headers, values))
        if row.get(key) not in (None, ""):
            yield row


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--detail-sheet", default="打标明细")
    parser.add_argument("--dictionary-sheet", default="标签字典")
    parser.add_argument("--json-out", type=Path)
    args = parser.parse_args()

    try:
        result = validate(args.workbook, args.detail_sheet, args.dictionary_sheet)
    except Exception as exc:
        result = {"status": "fail", "workbook": str(args.workbook), "errors": [str(exc)], "warnings": []}

    output = json.dumps(result, ensure_ascii=False, indent=2)
    print(output)
    if args.json_out:
        args.json_out.write_text(output, encoding="utf-8")
    return 0 if result["status"] == "pass" else 1


if __name__ == "__main__":
    sys.exit(main())
